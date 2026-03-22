import os

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo

from database.db import SessionLocal
from database.models import Invoice, LineItem


def fetch_data():

    session = SessionLocal()

    invoices = session.query(Invoice).all()
    line_items = session.query(LineItem).all()

    session.close()

    return invoices, line_items

def generate_report():

    invoices, line_items = fetch_data()

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Invoices"

    headers = [
        "Vendor",
        "Invoice Number",
        "Invoice Date",
        "Due Date",
        "Total Amount",
        "Category",
        "File Path"
    ]

    ws1.append(headers)
    table = Table(displayName="InvoiceTable", ref=f"A1:G{len(invoices)+1}")

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    table.tableStyleInfo = style
    ws1.add_table(table)

    for inv in invoices:

       ws1.append([
        inv.vendor_name or "Unknown Vendor",
        inv.invoice_number or "N/A",
        inv.invoice_date or "N/A",
        inv.due_date or "N/A",
        inv.total_amount or 0,
        inv.category or "Other",
        inv.file_path or ""
    ])

    ws2 = wb.create_sheet("Line Items")

    headers = [
        "Invoice ID",
        "Description",
        "Quantity",
        "Unit Price",
        "Item Total"
    ]

    ws2.append(headers)

    for item in line_items:

        ws2.append([
            item.invoice_id,
            item.description,
            item.quantity,
            item.unit_price,
            item.item_total
        ])

    ws3 = wb.create_sheet("Category Summary")

    category_totals = {}

    for inv in invoices:

        category = inv.category if inv.category else "Other"
        total = inv.total_amount if inv.total_amount else 0

        if category not in category_totals:
            category_totals[category] = 0

        category_totals[category] += total


    ws3.append(["Category", "Total Spend"])

    for category, total in category_totals.items():
        ws3.append([category, total])

    chart = PieChart()

    data = Reference(
        ws3,
        min_col=2,
        min_row=1,
        max_row=len(category_totals) + 1
    )

    labels = Reference(
        ws3,
        min_col=1,
        min_row=2,
        max_row=len(category_totals) + 1
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)

    chart.title = "Spend by Category"

    ws3.add_chart(chart, "E5")

    os.makedirs("output", exist_ok=True)
    from datetime import datetime
    date = datetime.now().strftime("%Y_%m")
    file_path = f"output/expense_report_{date}.xlsx"

    wb.save(file_path)

    print("Excel report generated:", file_path)

if __name__ == "__main__":

    generate_report()
